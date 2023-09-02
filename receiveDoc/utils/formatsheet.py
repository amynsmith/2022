#!/usr/bin/env python
# coding: utf-8
import pandas as pd
import re

# from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.formatting.rule import Rule
from openpyxl import load_workbook



def highlight_respdflist(fn):
    wb = load_workbook(fn)
    sheet=wb.active
    sheet.freeze_panes = "B2"
    sheet.column_dimensions['B'].width =30 #docmark
    sheet.column_dimensions['D'].width=80 #title

    # highlight body file
    for row in sheet.iter_rows(min_col=2, max_col=13, min_row=2, values_only=False):
        if row[1].value==False:
            # for cell in row[0:4]:
            for cell in row:
                cell.style="Headline 3"

    # highlight false values
    red_text = Font(color="9C0006")
    red_fill = PatternFill(bgColor="FFC7CE")
    # dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = CellIsRule(operator='equal', formula=['false'], font=red_text, fill=red_fill, stopIfTrue=True)
    sheet.conditional_formatting.add('L1:M100', rule)
    wb.save(fn)




if __name__ == "__main__":
    base_path = r"C:\Users\Amy19\Desktop\收文"
    basedir = base_path + pd.Timestamp.today().strftime("%Y%m%d")
    resfn = basedir + "\\" + "respdflist.xlsx"
    highlight_respdflist(resfn)
    input("...\n")