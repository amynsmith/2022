#!/usr/bin/env python
# coding: utf-8

import re

# from openpyxl import Workbook
# from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import Font
# from openpyxl.styles.differential import DifferentialStyle
# from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
# from openpyxl.formatting.rule import Rule
from openpyxl import load_workbook



def addformat(fn,resfn):
    wb = load_workbook(fn)
    sheet = wb.active
    sheet.freeze_panes = "D2"
    sheet.column_dimensions['D'].width = 30  # 备注列加宽

    for row in sheet.iter_rows(min_col=6, max_col=8, min_row=2, values_only=False):
        if (re.search("请示\Z", row[0].value) and (row[2].value == "发文")):
            row[0].style = "Accent1"
        elif (re.search("批复\Z", row[0].value) and (row[2].value == "收文")):
            row[0].style = "Accent5"
        elif (re.search("任(职|免)", row[0].value) and (row[2].value == "收文")):
            row[0].style = "Accent4"
        else:
            continue
        # 默认样式字号为12，需要将标题列字号重新改为11
        row[0].font = Font(size=11, color="FFFFFF")

    wb.save(resfn)



if __name__ == "__main__":
    # fn="res_sort.xlsx"
    # resfn="res_format.xlsx"

    fn="res_sort_beijing.xlsx"
    resfn="res_format_beijing.xlsx"
    addformat(fn,resfn)

